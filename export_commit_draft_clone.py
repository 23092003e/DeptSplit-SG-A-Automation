#!/usr/bin/env python3
"""
Export script that augments existing per-department files by adding Commit and Draft Commit sheets.
Preserves formatting using openpyxl in clone mode.
"""

import argparse
import logging
from pathlib import Path
from typing import List, Optional, Set

import pandas as pd
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def sanitize_filename(text: str) -> str:
    """Sanitize text for use as a filename by replacing invalid characters with dashes."""
    if not text or not text.strip():
        return "Unknown"
    
    # Replace invalid filename characters with dashes
    invalid_chars = r'[\\/:*?"<>|]'
    sanitized = re.sub(invalid_chars, '-', text.strip())
    
    # Remove multiple consecutive dashes and leading/trailing dashes
    sanitized = re.sub(r'-+', '-', sanitized).strip('-')
    
    return sanitized if sanitized else "Unknown"


def load_workbook_safe(path: Path) -> Workbook:
    """Safely load an Excel workbook."""
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")
    
    if not path.suffix.lower() in ['.xlsx', '.xlsm']:
        raise ValueError(f"File must be an Excel file (.xlsx or .xlsm): {path}")
    
    try:
        return openpyxl.load_workbook(path, data_only=False, read_only=False)
    except Exception as e:
        raise ValueError(f"Failed to load Excel file {path}: {e}")


def get_master_group_list(workbook: Workbook) -> Set[str]:
    """
    Read unique, non-null values from SG&A Summary → Project/Department column.
    
    Args:
        workbook: Source openpyxl workbook
        
    Returns:
        Set of unique group names
    """
    # Find SG&A Summary sheet
    sga_sheet = None
    for sheet_name in workbook.sheetnames:
        if 'SG&A Summary' in sheet_name or 'SGA Summary' in sheet_name:
            sga_sheet = workbook[sheet_name]
            break
    
    if sga_sheet is None:
        raise ValueError("SG&A Summary sheet not found in workbook")
    
    # Find the actual header row by searching for rows with Department/Project keywords
    header_row = None
    dept_col_idx = None
    
    # Search through first 10 rows to find header row
    for row_idx in range(1, min(11, sga_sheet.max_row + 1)):
        for col_idx in range(1, sga_sheet.max_column + 1):
            cell_value = sga_sheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                cell_str = str(cell_value).lower().strip()
                # Look for department/project column indicators
                if ('department' in cell_str or 'project' in cell_str or 
                    'dept' in cell_str or 'group' in cell_str):
                    header_row = row_idx
                    dept_col_idx = col_idx
                    logger.info(f"Found department/project column: '{cell_value}' at row {header_row}, column {col_idx}")
                    break
        if header_row is not None:
            break
    
    if header_row is None or dept_col_idx is None:
        # Debug: Print first few rows to help identify the structure
        logger.info("Searching for header row - first 5 rows of data:")
        for row_idx in range(1, min(6, sga_sheet.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(11, sga_sheet.max_column + 1)):  # First 10 columns
                cell_value = sga_sheet.cell(row=row_idx, column=col_idx).value
                row_data.append(f"'{cell_value}'")
            logger.info(f"Row {row_idx}: {'; '.join(row_data)}")
        raise ValueError("Project/Department column not found in SG&A Summary sheet")
    
    # Collect unique, non-null values
    groups = set()
    for row_idx in range(header_row + 1, sga_sheet.max_row + 1):
        cell_value = sga_sheet.cell(row=row_idx, column=dept_col_idx).value
        if cell_value and str(cell_value).strip():
            groups.add(str(cell_value).strip())
    
    logger.info(f"Found {len(groups)} unique groups in SG&A Summary: {sorted(groups)}")
    return groups


def remove_unnamed_columns(worksheet: Worksheet, split_col_name: str) -> None:
    """Remove columns whose header starts with 'Unnamed' but never remove the split column."""
    if worksheet.max_row == 0:
        return
    
    # Identify columns to remove (in reverse order to avoid index shifting)
    cols_to_remove = []
    for col_idx in range(worksheet.max_column, 0, -1):
        header_cell = worksheet.cell(row=1, column=col_idx)
        if header_cell.value:
            header_value = str(header_cell.value).strip()
            # Remove if starts with "Unnamed" (case-insensitive) but not the split column
            if (header_value.lower().startswith('unnamed') and 
                header_value != split_col_name):
                cols_to_remove.append(col_idx)
    
    # Remove columns
    for col_idx in cols_to_remove:
        worksheet.delete_cols(col_idx)
        logger.info(f"Removed column {col_idx} (Unnamed column)")


def copy_cell_format(source_cell, target_cell):
    """Copy formatting from source cell to target cell."""
    if source_cell.font:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            vertAlign=source_cell.font.vertAlign,
            underline=source_cell.font.underline,
            strike=source_cell.font.strike,
            color=source_cell.font.color
        )
    
    if source_cell.fill:
        target_cell.fill = PatternFill(
            fill_type=source_cell.fill.fill_type,
            start_color=source_cell.fill.start_color,
            end_color=source_cell.fill.end_color
        )
    
    if source_cell.border:
        target_cell.border = Border(
            left=source_cell.border.left,
            right=source_cell.border.right,
            top=source_cell.border.top,
            bottom=source_cell.border.bottom,
            diagonal=source_cell.border.diagonal,
            diagonal_direction=source_cell.border.diagonal_direction,
            outline=source_cell.border.outline,
            vertical=source_cell.border.vertical,
            horizontal=source_cell.border.horizontal
        )
    
    if source_cell.alignment:
        target_cell.alignment = Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            text_rotation=source_cell.alignment.text_rotation,
            wrap_text=source_cell.alignment.wrap_text,
            shrink_to_fit=source_cell.alignment.shrink_to_fit,
            indent=source_cell.alignment.indent
        )
    
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format


def copy_column_widths(source_ws: Worksheet, target_ws: Worksheet):
    """Copy column widths from source to target worksheet."""
    for col_letter in source_ws.column_dimensions:
        if col_letter in source_ws.column_dimensions:
            target_ws.column_dimensions[col_letter].width = source_ws.column_dimensions[col_letter].width


def copy_row_heights(source_ws: Worksheet, target_ws: Worksheet, max_rows: int):
    """Copy row heights from source to target worksheet."""
    for row_num in range(1, max_rows + 1):
        if row_num in source_ws.row_dimensions:
            target_ws.row_dimensions[row_num].height = source_ws.row_dimensions[row_num].height


def process_sheet_clone_mode(source_workbook: Workbook, sheet_name: str, split_col: str, 
                            group: str, target_workbook: Workbook) -> int:
    """
    Process SG&A Summary sheet in CLONE mode - copy entire sheet then filter by deleting non-matching rows.
    
    Args:
        source_workbook: Source openpyxl workbook
        sheet_name: Name of sheet to process
        split_col: Column name to split by
        group: Group value to filter by
        target_workbook: Target openpyxl workbook
        
    Returns:
        Number of rows remaining after filtering
    """
    if sheet_name not in source_workbook.sheetnames:
        logger.warning(f"Sheet '{sheet_name}' not found in source workbook")
        return 0
    
    source_ws = source_workbook[sheet_name]
    
    # Create or get target worksheet
    if sheet_name in target_workbook.sheetnames:
        target_ws = target_workbook[sheet_name]
        target_workbook.remove(target_ws)
    
    target_ws = target_workbook.create_sheet(title=sheet_name)
    
    # Step 1: Clone the entire sheet with all formatting
    for row_num in range(1, source_ws.max_row + 1):
        for col_num in range(1, source_ws.max_column + 1):
            source_cell = source_ws.cell(row=row_num, column=col_num)
            target_cell = target_ws.cell(row=row_num, column=col_num)
            
            # Copy value and formatting
            target_cell.value = source_cell.value
            copy_cell_format(source_cell, target_cell)
    
    # Copy column widths and row heights
    copy_column_widths(source_ws, target_ws)
    copy_row_heights(source_ws, target_ws, source_ws.max_row)
    
    # Step 2: Find the split column index
    split_col_idx = None
    header_row = None
    
    # Find header row and split column (similar to get_master_group_list logic)
    for row_idx in range(1, min(11, target_ws.max_row + 1)):
        for col_idx in range(1, target_ws.max_column + 1):
            cell_value = target_ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                cell_str = str(cell_value).lower().strip()
                if split_col.lower() in cell_str or (split_col.lower() == "project/department" and 
                    ('department' in cell_str or 'project' in cell_str)):
                    header_row = row_idx
                    split_col_idx = col_idx
                    logger.info(f"Found split column: '{cell_value}' at row {header_row}, column {col_idx} in sheet '{sheet_name}'")
                    break
        if header_row is not None:
            break
    
    if split_col_idx is None or header_row is None:
        logger.warning(f"Split column '{split_col}' not found in sheet '{sheet_name}' - keeping all rows")
        rows_remaining = target_ws.max_row - 1 if target_ws.max_row > 1 else 0
    else:
        # Step 3: Delete rows that don't match the group (in reverse order to avoid index shifting)
        rows_to_delete = []
        for row_idx in range(header_row + 1, target_ws.max_row + 1):
            cell_value = target_ws.cell(row=row_idx, column=split_col_idx).value
            if cell_value is None or str(cell_value).strip() != group:
                rows_to_delete.append(row_idx)
        
        # Delete rows in reverse order
        for row_idx in reversed(rows_to_delete):
            target_ws.delete_rows(row_idx)
        
        rows_remaining = target_ws.max_row - header_row if target_ws.max_row > header_row else 0
        logger.info(f"Deleted {len(rows_to_delete)} non-matching rows, {rows_remaining} rows remaining")
    
    # Apply AutoFilter if there's data
    if target_ws.max_row > 1 and target_ws.max_column > 0:
        target_ws.auto_filter = AutoFilter(
            ref=f"A1:{get_column_letter(target_ws.max_column)}{target_ws.max_row}"
        )
    
    logger.info(f"Sheet '{sheet_name}' filtered for group '{group}': {rows_remaining} rows remaining")
    return rows_remaining


def process_sheet(source_workbook: Workbook, sheet_name: str, split_col: str, 
                 group: str, target_workbook: Workbook) -> int:
    """
    Process a single sheet for a specific group.
    
    Args:
        source_workbook: Source openpyxl workbook
        sheet_name: Name of sheet to process
        split_col: Column name to split by
        group: Group value to filter by
        target_workbook: Target openpyxl workbook
        
    Returns:
        Number of rows copied
    """
    if sheet_name not in source_workbook.sheetnames:
        logger.warning(f"Sheet '{sheet_name}' not found in source workbook")
        return 0
    
    source_ws = source_workbook[sheet_name]
    
    # Find the split column index with flexible matching
    split_col_idx = None
    header_row = 1  # 1-based indexing for most sheets
    
    # First try exact match
    for col_idx in range(1, source_ws.max_column + 1):
        cell_value = source_ws.cell(row=header_row, column=col_idx).value
        if cell_value and str(cell_value).strip() == split_col:
            split_col_idx = col_idx
            break
    
    # If not found, try partial matching for "Department"
    if split_col_idx is None and "department" in split_col.lower():
        for col_idx in range(1, source_ws.max_column + 1):
            cell_value = source_ws.cell(row=header_row, column=col_idx).value
            if cell_value and "department" in str(cell_value).lower():
                split_col_idx = col_idx
                logger.info(f"Found department column: '{cell_value}' at column {col_idx} in sheet '{sheet_name}'")
                break
    
    if split_col_idx is None:
        logger.warning(f"Split column '{split_col}' not found in sheet '{sheet_name}'")
        return 0
    
    # Create or get target worksheet
    if sheet_name in target_workbook.sheetnames:
        target_ws = target_workbook[sheet_name]
        # Clear existing content
        target_workbook.remove(target_ws)
    
    target_ws = target_workbook.create_sheet(title=sheet_name)
    
    # Copy header row first
    for col_idx in range(1, source_ws.max_column + 1):
        source_cell = source_ws.cell(row=header_row, column=col_idx)
        target_cell = target_ws.cell(row=1, column=col_idx)
        target_cell.value = source_cell.value
        copy_cell_format(source_cell, target_cell)
    
    # Copy matching rows
    rows_copied = 0
    target_row = 2  # Start after header
    
    for source_row in range(header_row + 1, source_ws.max_row + 1):
        dept_cell = source_ws.cell(row=source_row, column=split_col_idx)
        if dept_cell.value and str(dept_cell.value).strip() == group:
            # Copy entire row
            for col_idx in range(1, source_ws.max_column + 1):
                source_cell = source_ws.cell(row=source_row, column=col_idx)
                target_cell = target_ws.cell(row=target_row, column=col_idx)
                target_cell.value = source_cell.value
                copy_cell_format(source_cell, target_cell)
            
            rows_copied += 1
            target_row += 1
    
    if rows_copied > 0:
        # Copy column widths and row heights
        copy_column_widths(source_ws, target_ws)
        copy_row_heights(source_ws, target_ws, target_row - 1)
        
        # Remove unnamed columns
        remove_unnamed_columns(target_ws, split_col)
        
        # Apply AutoFilter
        if target_ws.max_row > 1 and target_ws.max_column > 0:
            target_ws.auto_filter = AutoFilter(
                ref=f"A1:{get_column_letter(target_ws.max_column)}{target_ws.max_row}"
            )
    
    logger.info(f"Sheet '{sheet_name}' for group '{group}': {rows_copied} rows copied")
    return rows_copied


def main():
    parser = argparse.ArgumentParser(
        description="Augment existing per-department files by adding Commit and Draft Commit sheets"
    )
    parser.add_argument("--input", "-i", required=True, type=Path, 
                       help="Path to input Excel file")
    parser.add_argument("--outdir", "-o", required=True, type=Path,
                       help="Output directory")
    
    args = parser.parse_args()
    
    try:
        # Validate inputs
        input_path = Path(args.input)
        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found: {input_path}")
        
        output_dir = Path(args.outdir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Load source workbook
        logger.info(f"Loading workbook: {input_path}")
        source_wb = load_workbook_safe(input_path)
        
        # Get master group list
        logger.info("Reading master group list from SG&A Summary sheet")
        groups = get_master_group_list(source_wb)
        
        # Sheets to process - thêm SG&A Summary sheet với CLONE mode
        sheets_to_process = [
            {"name": "SG&A Summary", "header_row": 0, "split_col": "Project/Department", "clone_mode": True},
            {"name": "Commit", "header_row": 0, "split_col": "Department", "clone_mode": False},
            {"name": "Draft Commit", "header_row": 0, "split_col": "Department", "clone_mode": False}
        ]
        
        # Process each group
        for group in sorted(groups):
            logger.info(f"Processing group: {group}")
            
            # Generate target filename
            sanitized_name = sanitize_filename(group)
            target_filename = f"{sanitized_name}_SG&A Budget report.xlsx"
            target_path = output_dir / target_filename
            
            # Load or create target workbook
            if target_path.exists():
                logger.info(f"Opening existing file: {target_path}")
                target_wb = load_workbook_safe(target_path)
            else:
                logger.info(f"Creating new file: {target_path}")
                target_wb = Workbook()
                # Remove default sheet
                if 'Sheet' in target_wb.sheetnames:
                    target_wb.remove(target_wb['Sheet'])
            
            # Process each sheet
            total_rows = 0
            for sheet_config in sheets_to_process:
                sheet_name = sheet_config["name"]
                split_col = sheet_config["split_col"]
                clone_mode = sheet_config.get("clone_mode", False)
                
                if clone_mode:
                    # Use clone mode for SG&A Summary (copy entire sheet)
                    rows_copied = process_sheet_clone_mode(
                        source_wb, sheet_name, split_col, group, target_wb
                    )
                else:
                    # Use normal filtered mode for Commit/Draft Commit
                    rows_copied = process_sheet(
                        source_wb, sheet_name, split_col, group, target_wb
                    )
                total_rows += rows_copied
            
            # Save target workbook if any rows were copied
            if total_rows > 0:
                target_wb.save(target_path)
                logger.info(f"Saved {target_path} with {total_rows} total rows")
            else:
                logger.warning(f"No data found for group '{group}', skipping file creation")
                if not target_path.exists():
                    # Don't create empty files
                    pass
        
        logger.info("Processing completed successfully")
        
    except Exception as e:
        logger.error(f"Error: {e}")
        return 1
    
    return 0


if __name__ == "__main__":
    import re  # Import needed for sanitize_filename
    exit(main())