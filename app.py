#!/usr/bin/env python3
"""
Streamlit app for Budget Report Processor
Splits Excel budget files by department/group with preserved formatting
"""

import streamlit as st
import pandas as pd
import openpyxl
from pathlib import Path
import tempfile
import zipfile
import io
import re
import logging
from typing import List, Optional, Set

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ===== UTILITY FUNCTIONS FROM ORIGINAL SCRIPT =====

def sanitize_filename(text: str) -> str:
    """Sanitize text for use as a filename by replacing invalid characters with dashes."""
    if not text or not text.strip():
        return "Unknown"
    
    # Replace invalid filename characters with dashes
    invalid_chars = r'[\\/:*?"<>|]'
    sanitized = re.sub(invalid_chars, '-', text.strip())
    
    # Remove multiple consecutive dashes and leading/trailing dashes
    sanitized = re.sub(r'-+', '-', sanitized).strip('-')
    
    return sanitized if sanitized else "Unknown"

def load_workbook_safe(path: Path) -> Workbook:
    """Safely load an Excel workbook."""
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")
    
    if not path.suffix.lower() in ['.xlsx', '.xlsm']:
        raise ValueError(f"File must be an Excel file (.xlsx or .xlsm): {path}")
    
    try:
        return openpyxl.load_workbook(path, data_only=False, read_only=False)
    except Exception as e:
        raise ValueError(f"Failed to load Excel file {path}: {e}")

def get_master_group_list(workbook: Workbook) -> Set[str]:
    """Read unique, non-null values from SG&A Summary → Project/Department column."""
    # Find SG&A Summary sheet
    sga_sheet = None
    for sheet_name in workbook.sheetnames:
        if 'SG&A Summary' in sheet_name or 'SGA Summary' in sheet_name:
            sga_sheet = workbook[sheet_name]
            break
    
    if sga_sheet is None:
        raise ValueError("SG&A Summary sheet not found in workbook")
    
    # Find the actual header row by searching for rows with Department/Project keywords
    header_row = None
    dept_col_idx = None
    
    # Search through first 10 rows to find header row
    for row_idx in range(1, min(11, sga_sheet.max_row + 1)):
        for col_idx in range(1, sga_sheet.max_column + 1):
            cell_value = sga_sheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                cell_str = str(cell_value).lower().strip()
                # Look for department/project column indicators
                if ('department' in cell_str or 'project' in cell_str or 
                    'dept' in cell_str or 'group' in cell_str):
                    header_row = row_idx
                    dept_col_idx = col_idx
                    logger.info(f"Found department/project column: '{cell_value}' at row {header_row}, column {col_idx}")
                    break
        if header_row is not None:
            break
    
    if header_row is None or dept_col_idx is None:
        raise ValueError("Project/Department column not found in SG&A Summary sheet")
    
    # Collect unique, non-null values
    groups = set()
    for row_idx in range(header_row + 1, sga_sheet.max_row + 1):
        cell_value = sga_sheet.cell(row=row_idx, column=dept_col_idx).value
        if cell_value and str(cell_value).strip():
            groups.add(str(cell_value).strip())
    
    logger.info(f"Found {len(groups)} unique groups in SG&A Summary: {sorted(groups)}")
    return groups

def copy_cell_format(source_cell, target_cell):
    """Copy formatting from source cell to target cell."""
    if source_cell.font:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            vertAlign=source_cell.font.vertAlign,
            underline=source_cell.font.underline,
            strike=source_cell.font.strike,
            color=source_cell.font.color
        )
    
    if source_cell.fill:
        target_cell.fill = PatternFill(
            fill_type=source_cell.fill.fill_type,
            start_color=source_cell.fill.start_color,
            end_color=source_cell.fill.end_color
        )
    
    if source_cell.border:
        target_cell.border = Border(
            left=source_cell.border.left,
            right=source_cell.border.right,
            top=source_cell.border.top,
            bottom=source_cell.border.bottom,
            diagonal=source_cell.border.diagonal,
            diagonal_direction=source_cell.border.diagonal_direction,
            outline=source_cell.border.outline,
            vertical=source_cell.border.vertical,
            horizontal=source_cell.border.horizontal
        )
    
    if source_cell.alignment:
        target_cell.alignment = Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            text_rotation=source_cell.alignment.text_rotation,
            wrap_text=source_cell.alignment.wrap_text,
            shrink_to_fit=source_cell.alignment.shrink_to_fit,
            indent=source_cell.alignment.indent
        )
    
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format

def copy_column_widths(source_ws: Worksheet, target_ws: Worksheet):
    """Copy column widths from source to target worksheet."""
    for col_letter in source_ws.column_dimensions:
        if col_letter in source_ws.column_dimensions:
            target_ws.column_dimensions[col_letter].width = source_ws.column_dimensions[col_letter].width

def copy_row_heights(source_ws: Worksheet, target_ws: Worksheet, max_rows: int):
    """Copy row heights from source to target worksheet."""
    for row_num in range(1, max_rows + 1):
        if row_num in source_ws.row_dimensions:
            target_ws.row_dimensions[row_num].height = source_ws.row_dimensions[row_num].height

def remove_unnamed_columns(worksheet: Worksheet, split_col_name: str) -> None:
    """Remove columns whose header starts with 'Unnamed' but never remove the split column."""
    if worksheet.max_row == 0:
        return
    
    # Identify columns to remove (in reverse order to avoid index shifting)
    cols_to_remove = []
    for col_idx in range(worksheet.max_column, 0, -1):
        header_cell = worksheet.cell(row=1, column=col_idx)
        if header_cell.value:
            header_value = str(header_cell.value).strip()
            # Remove if starts with "Unnamed" (case-insensitive) but not the split column
            if (header_value.lower().startswith('unnamed') and 
                header_value != split_col_name):
                cols_to_remove.append(col_idx)
    
    # Remove columns
    for col_idx in cols_to_remove:
        worksheet.delete_cols(col_idx)
        logger.info(f"Removed column {col_idx} (Unnamed column)")

def process_sheet_clone_mode(source_workbook: Workbook, sheet_name: str, split_col: str, 
                            group: str, target_workbook: Workbook) -> int:
    """Process SG&A Summary sheet in CLONE mode - copy entire sheet then filter by deleting non-matching rows."""
    if sheet_name not in source_workbook.sheetnames:
        logger.warning(f"Sheet '{sheet_name}' not found in source workbook")
        return 0
    
    source_ws = source_workbook[sheet_name]
    
    # Create or get target worksheet
    if sheet_name in target_workbook.sheetnames:
        target_ws = target_workbook[sheet_name]
        target_workbook.remove(target_ws)
    
    target_ws = target_workbook.create_sheet(title=sheet_name)
    
    # Step 1: Clone the entire sheet with all formatting
    for row_num in range(1, source_ws.max_row + 1):
        for col_num in range(1, source_ws.max_column + 1):
            source_cell = source_ws.cell(row=row_num, column=col_num)
            target_cell = target_ws.cell(row=row_num, column=col_num)
            
            # Copy value and formatting
            target_cell.value = source_cell.value
            copy_cell_format(source_cell, target_cell)
    
    # Copy column widths and row heights
    copy_column_widths(source_ws, target_ws)
    copy_row_heights(source_ws, target_ws, source_ws.max_row)
    
    # Step 2: Find the split column index
    split_col_idx = None
    header_row = None
    
    # Find header row and split column
    for row_idx in range(1, min(11, target_ws.max_row + 1)):
        for col_idx in range(1, target_ws.max_column + 1):
            cell_value = target_ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                cell_str = str(cell_value).lower().strip()
                if split_col.lower() in cell_str or (split_col.lower() == "project/department" and 
                    ('department' in cell_str or 'project' in cell_str)):
                    header_row = row_idx
                    split_col_idx = col_idx
                    logger.info(f"Found split column: '{cell_value}' at row {header_row}, column {col_idx} in sheet '{sheet_name}'")
                    break
        if header_row is not None:
            break
    
    if split_col_idx is None or header_row is None:
        logger.warning(f"Split column '{split_col}' not found in sheet '{sheet_name}' - keeping all rows")
        rows_remaining = target_ws.max_row - 1 if target_ws.max_row > 1 else 0
    else:
        # Step 3: Delete rows that don't match the group (in reverse order to avoid index shifting)
        rows_to_delete = []
        for row_idx in range(header_row + 1, target_ws.max_row + 1):
            cell_value = target_ws.cell(row=row_idx, column=split_col_idx).value
            if cell_value is None or str(cell_value).strip() != group:
                rows_to_delete.append(row_idx)
        
        # Delete rows in reverse order
        for row_idx in reversed(rows_to_delete):
            target_ws.delete_rows(row_idx)
        
        rows_remaining = target_ws.max_row - header_row if target_ws.max_row > header_row else 0
        logger.info(f"Deleted {len(rows_to_delete)} non-matching rows, {rows_remaining} rows remaining")
    
    # Apply AutoFilter if there's data
    if target_ws.max_row > 1 and target_ws.max_column > 0:
        target_ws.auto_filter = AutoFilter(
            ref=f"A1:{get_column_letter(target_ws.max_column)}{target_ws.max_row}"
        )
    
    logger.info(f"Sheet '{sheet_name}' filtered for group '{group}': {rows_remaining} rows remaining")
    return rows_remaining

def process_sheet(source_workbook: Workbook, sheet_name: str, split_col: str, 
                 group: str, target_workbook: Workbook) -> int:
    """Process a single sheet for a specific group."""
    if sheet_name not in source_workbook.sheetnames:
        logger.warning(f"Sheet '{sheet_name}' not found in source workbook")
        return 0
    
    source_ws = source_workbook[sheet_name]
    
    # Find the split column index with flexible matching
    split_col_idx = None
    header_row = 1  # 1-based indexing for most sheets
    
    # First try exact match
    for col_idx in range(1, source_ws.max_column + 1):
        cell_value = source_ws.cell(row=header_row, column=col_idx).value
        if cell_value and str(cell_value).strip() == split_col:
            split_col_idx = col_idx
            break
    
    # If not found, try partial matching for "Department"
    if split_col_idx is None and "department" in split_col.lower():
        for col_idx in range(1, source_ws.max_column + 1):
            cell_value = source_ws.cell(row=header_row, column=col_idx).value
            if cell_value and "department" in str(cell_value).lower():
                split_col_idx = col_idx
                logger.info(f"Found department column: '{cell_value}' at column {col_idx} in sheet '{sheet_name}'")
                break
    
    if split_col_idx is None:
        logger.warning(f"Split column '{split_col}' not found in sheet '{sheet_name}'")
        return 0
    
    # Create or get target worksheet
    if sheet_name in target_workbook.sheetnames:
        target_ws = target_workbook[sheet_name]
        target_workbook.remove(target_ws)
    
    target_ws = target_workbook.create_sheet(title=sheet_name)
    
    # Copy header row first
    for col_idx in range(1, source_ws.max_column + 1):
        source_cell = source_ws.cell(row=header_row, column=col_idx)
        target_cell = target_ws.cell(row=1, column=col_idx)
        target_cell.value = source_cell.value
        copy_cell_format(source_cell, target_cell)
    
    # Copy matching rows
    rows_copied = 0
    target_row = 2  # Start after header
    
    for source_row in range(header_row + 1, source_ws.max_row + 1):
        dept_cell = source_ws.cell(row=source_row, column=split_col_idx)
        if dept_cell.value and str(dept_cell.value).strip() == group:
            # Copy entire row
            for col_idx in range(1, source_ws.max_column + 1):
                source_cell = source_ws.cell(row=source_row, column=col_idx)
                target_cell = target_ws.cell(row=target_row, column=col_idx)
                target_cell.value = source_cell.value
                copy_cell_format(source_cell, target_cell)
            
            rows_copied += 1
            target_row += 1
    
    if rows_copied > 0:
        # Copy column widths and row heights
        copy_column_widths(source_ws, target_ws)
        copy_row_heights(source_ws, target_ws, target_row - 1)
        
        # Remove unnamed columns
        remove_unnamed_columns(target_ws, split_col)
        
        # Apply AutoFilter
        if target_ws.max_row > 1 and target_ws.max_column > 0:
            target_ws.auto_filter = AutoFilter(
                ref=f"A1:{get_column_letter(target_ws.max_column)}{target_ws.max_row}"
            )
    
    logger.info(f"Sheet '{sheet_name}' for group '{group}': {rows_copied} rows copied")
    return rows_copied

# ===== STREAMLIT APP FUNCTIONS =====

def process_budget_files(uploaded_file, progress_callback=None):
    """Process the uploaded budget file and return processed files."""
    
    # Create temporary directory for processing
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        
        # Save uploaded file
        input_path = temp_path / "input.xlsx"
        with open(input_path, "wb") as f:
            f.write(uploaded_file.getvalue())
        
        # Load source workbook
        if progress_callback:
            progress_callback("Loading workbook...")
        source_wb = load_workbook_safe(input_path)
        
        # Get master group list
        if progress_callback:
            progress_callback("Analyzing groups...")
        groups = get_master_group_list(source_wb)
        
        # Create output directory
        output_dir = temp_path / "output"
        output_dir.mkdir(exist_ok=True)
        
        # Sheets to process
        sheets_to_process = [
            {"name": "SG&A Summary", "split_col": "Project/Department", "clone_mode": True},
            {"name": "Commit", "split_col": "Department", "clone_mode": False},
            {"name": "Draft Commit", "split_col": "Department", "clone_mode": False}
        ]
        
        processed_files = []
        total_groups = len(groups)
        
        # Process each group
        for i, group in enumerate(sorted(groups)):
            if progress_callback:
                progress_callback(f"Processing group {i+1}/{total_groups}: {group}")
            
            # Generate target filename
            sanitized_name = sanitize_filename(group)
            target_filename = f"{sanitized_name}_SG&A Budget report.xlsx"
            target_path = output_dir / target_filename
            
            # Create target workbook
            target_wb = Workbook()
            if 'Sheet' in target_wb.sheetnames:
                target_wb.remove(target_wb['Sheet'])
            
            # Process each sheet
            total_rows = 0
            for sheet_config in sheets_to_process:
                sheet_name = sheet_config["name"]
                split_col = sheet_config["split_col"]
                clone_mode = sheet_config.get("clone_mode", False)
                
                if clone_mode:
                    rows_copied = process_sheet_clone_mode(
                        source_wb, sheet_name, split_col, group, target_wb
                    )
                else:
                    rows_copied = process_sheet(
                        source_wb, sheet_name, split_col, group, target_wb
                    )
                total_rows += rows_copied
            
            # Save target workbook if any rows were copied
            if total_rows > 0:
                target_wb.save(target_path)
                processed_files.append((target_filename, target_path))
        
        # Read all files into memory before temp directory is deleted
        if progress_callback:
            progress_callback("Preparing files for download...")
        
        files_in_memory = []
        for filename, filepath in processed_files:
            with open(filepath, 'rb') as f:
                file_data = f.read()
            files_in_memory.append((filename, file_data))
        
        return files_in_memory

def create_zip_download(files_in_memory):
    """Create a ZIP file containing all processed files from memory."""
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for filename, file_data in files_in_memory:
            zip_file.writestr(filename, file_data)
    
    zip_buffer.seek(0)
    return zip_buffer

# ===== STREAMLIT APP MAIN FUNCTION =====

def main():
    st.set_page_config(
        page_title="Budget Report Processor",
        page_icon="📊",
        layout="wide"
    )
    
    st.title("📊 SG&A Budget Report Processor")
    st.markdown("**Split your budget reports by department/group with preserved formatting**")
    
    # Sidebar for instructions
    with st.sidebar:
        st.header("ℹ️ Instructions")
        st.markdown("""
        1. Upload your Excel budget file
        2. The app will automatically detect departments/groups
        3. Each group will get its own Excel file with:
           - SG&A Summary (filtered)
           - Commit sheet (if exists)
           - Draft Commit sheet (if exists)
        4. Download all files as a ZIP
        
        **Requirements:**
        - Excel file (.xlsx or .xlsm)
        - Must contain "SG&A Summary" sheet
        - Project/Department column in SG&A Summary
        """)
        
        st.header("🔧 Sheet Configuration")
        st.info("""
        **SG&A Summary**: Clone mode (preserves all formatting)
        **Commit/Draft Commit**: Filter mode (copies matching rows)
        """)
    
    # Main content area
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.header("📁 File Upload")
        uploaded_file = st.file_uploader(
            "Choose your Excel budget file",
            type=['xlsx', 'xlsm'],
            help="Upload the Excel file containing your budget data"
        )
        
        if uploaded_file is not None:
            # File info
            st.success(f"✅ File uploaded: {uploaded_file.name}")
            st.info(f"📊 File size: {uploaded_file.size / 1024 / 1024:.2f} MB")
            
            # Preview option
            if st.checkbox("🔍 Preview file structure"):
                try:
                    # Create temporary file for preview
                    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                        tmp.write(uploaded_file.getvalue())
                        tmp_path = Path(tmp.name)
                    
                    # Load and show sheet names
                    wb = load_workbook_safe(tmp_path)
                    st.write("**Available sheets:**")
                    for sheet in wb.sheetnames:
                        st.write(f"- {sheet}")
                    
                    # Preview groups if possible
                    try:
                        groups = get_master_group_list(wb)
                        st.write("**Detected groups/departments:**")
                        for group in sorted(groups):
                            st.write(f"- {group}")
                        st.info(f"Total groups found: {len(groups)}")
                    except Exception as e:
                        st.warning(f"Could not detect groups: {e}")
                    
                    # Clean up
                    tmp_path.unlink()
                    
                except Exception as e:
                    st.error(f"Error previewing file: {e}")
    
    with col2:
        st.header("⚙️ Processing Options")
        
        # Processing settings
        preserve_formatting = st.checkbox("🎨 Preserve formatting", value=True, disabled=True)
        add_filters = st.checkbox("🔍 Add AutoFilters", value=True, disabled=True)
        remove_unnamed = st.checkbox("🧹 Remove unnamed columns", value=True, disabled=True)
        
        st.markdown("---")
        
        # Processing button
        if uploaded_file is not None:
            if st.button("🚀 Process Budget Files", type="primary", use_container_width=True):
                # Progress tracking
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                def update_progress(message):
                    status_text.text(message)
                
                try:
                    # Process files
                    with st.spinner("Processing budget files..."):
                        files_in_memory = process_budget_files(
                            uploaded_file, 
                            progress_callback=update_progress
                        )
                    
                    progress_bar.progress(100)
                    status_text.text("✅ Processing completed!")
                    
                    if files_in_memory:
                        st.success(f"🎉 Successfully processed {len(files_in_memory)} files!")
                        
                        # Show processed files
                        st.write("**Generated files:**")
                        for filename, _ in files_in_memory:
                            st.write(f"📄 {filename}")
                        
                        # Create and offer ZIP download
                        zip_buffer = create_zip_download(files_in_memory)
                        
                        st.download_button(
                            label="📥 Download All Files (ZIP)",
                            data=zip_buffer.getvalue(),
                            file_name="budget_reports_by_department.zip",
                            mime="application/zip",
                            type="primary",
                            use_container_width=True
                        )
                        
                    else:
                        st.warning("⚠️ No files were generated. Please check your input file structure.")
                
                except Exception as e:
                    st.error(f"❌ Error processing files: {e}")
                    st.exception(e)
        else:
            st.info("👆 Upload a file to start processing")
    
    # Footer
    st.markdown("---")
    st.markdown(
        """
        <div style='text-align: center; color: #666;'>
        Budget Report Processor v1.0 | Built with Streamlit
        </div>
        """, 
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()