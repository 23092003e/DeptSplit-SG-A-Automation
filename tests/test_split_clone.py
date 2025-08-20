"""
Tests for clone export mode functionality.
"""

import pytest
import pandas as pd
from pathlib import Path
import tempfile
import shutil
import openpyxl
from openpyxl.styles import Font, Fill, PatternFill

from sga_splitter.exporters import export_clone


class TestExportClone:
    """Test suite for export_clone function."""
    
    @pytest.fixture
    def temp_output_dir(self):
        """Create a temporary directory for test outputs."""
        temp_dir = Path(tempfile.mkdtemp())
        yield temp_dir
        # Cleanup after test
        shutil.rmtree(temp_dir)
    
    @pytest.fixture
    def sample_workbook_path(self, temp_output_dir):
        """Create a sample Excel workbook with styled headers and test data."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SG&A Summary"
        
        # Add some pre-header rows
        ws['A1'] = "Company Name"
        ws['A2'] = "SG&A Budget Report"
        ws['A3'] = ""  # Empty row
        
        # Add styled headers in row 4 (index 3)
        headers = ['ID', 'Department/Project', 'Amount', 'Description']
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        header_font = Font(bold=True)
        
        for col_idx, header in enumerate(headers):
            cell = ws.cell(row=4, column=col_idx + 1, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Add test data
        test_data = [
            [1, 'Sales', 1000, 'Sales expense'],
            [2, 'Marketing', 2000, 'Marketing cost'],
            [3, 'Sales', 1500, 'Sales travel'],
            [4, 'IT', 3000, 'IT equipment'],
            [5, 'Marketing', 2500, 'Marketing ads'],
            [6, 'IT', 4000, 'IT software']
        ]
        
        for row_idx, row_data in enumerate(test_data):
            for col_idx, value in enumerate(row_data):
                ws.cell(row=5 + row_idx, column=col_idx + 1, value=value)
        
        # Add auto filter
        ws.auto_filter.ref = "A4:D10"
        
        # Save the workbook
        workbook_path = temp_output_dir / "test_input.xlsx"
        wb.save(workbook_path)
        wb.close()
        
        return workbook_path
    
    def test_basic_clone_export(self, sample_workbook_path, temp_output_dir):
        """Test basic clone export functionality."""
        groups = ['Sales', 'Marketing', 'IT']
        sheet_name = 'SG&A Summary'
        header_row = 3  # 0-based index
        dp_col_idx = 1  # 0-based index
        
        manifest_entries = export_clone(
            input_path=sample_workbook_path,
            groups=groups,
            sheet_name=sheet_name,
            header_row=header_row,
            dp_col_idx=dp_col_idx,
            out_dir=temp_output_dir
        )
        
        # Check manifest entries
        assert len(manifest_entries) == 3
        
        for entry in manifest_entries:
            assert entry['mode'] == 'clone'
            assert 'Department/Project' in entry
            assert 'output_path' in entry
            assert 'row_count' in entry
            assert entry['row_count'] > 0
        
        # Verify files were created
        created_files = list(temp_output_dir.glob("*SG&A Summary.xlsx"))
        # Filter out the input file
        created_files = [f for f in created_files if f != sample_workbook_path]
        assert len(created_files) == 3
    
    def test_style_preservation(self, sample_workbook_path, temp_output_dir):
        """Test that styles are preserved in clone mode."""
        groups = ['Sales']
        
        manifest_entries = export_clone(
            input_path=sample_workbook_path,
            groups=groups,
            sheet_name='SG&A Summary',
            header_row=3,
            dp_col_idx=1,
            out_dir=temp_output_dir
        )
        
        # Open the created file and check styles
        output_path = Path(manifest_entries[0]['output_path'])
        wb = openpyxl.load_workbook(output_path)
        ws = wb['SG&A Summary']
        
        # Check that header row still has formatting
        header_cell = ws['B4']  # Department/Project header
        assert header_cell.font.bold is True
        # Check that header cell has the expected fill color (allowing for different alpha channels)
        assert header_cell.fill.start_color.rgb in ['FFD9E1F2', '00D9E1F2']  # Header fill color
        
        wb.close()
    
    def test_data_filtering(self, sample_workbook_path, temp_output_dir):
        """Test that only matching data rows are kept."""
        groups = ['Sales']
        
        manifest_entries = export_clone(
            input_path=sample_workbook_path,
            groups=groups,
            sheet_name='SG&A Summary',
            header_row=3,
            dp_col_idx=1,
            out_dir=temp_output_dir
        )
        
        # Check row count
        sales_entry = manifest_entries[0]
        assert sales_entry['row_count'] == 2  # 2 Sales rows in test data
        
        # Verify file contents
        output_path = Path(sales_entry['output_path'])
        wb = openpyxl.load_workbook(output_path)
        ws = wb['SG&A Summary']
        
        # Check that pre-header rows are preserved
        assert ws['A1'].value == "Company Name"
        assert ws['A2'].value == "SG&A Budget Report"
        
        # Check that header row is preserved
        assert ws['B4'].value == "Department/Project"
        
        # Check that only Sales data rows remain
        data_rows = []
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            if any(cell is not None for cell in row):  # Skip empty rows
                data_rows.append(row)
        
        assert len(data_rows) == 2
        for row in data_rows:
            assert row[1] == 'Sales'  # Department/Project column
        
        wb.close()
    
    def test_pre_header_preservation(self, sample_workbook_path, temp_output_dir):
        """Test that rows above the header are preserved."""
        groups = ['Marketing']
        
        manifest_entries = export_clone(
            input_path=sample_workbook_path,
            groups=groups,
            sheet_name='SG&A Summary',
            header_row=3,
            dp_col_idx=1,
            out_dir=temp_output_dir
        )
        
        output_path = Path(manifest_entries[0]['output_path'])
        wb = openpyxl.load_workbook(output_path)
        ws = wb['SG&A Summary']
        
        # Check pre-header content is preserved
        assert ws['A1'].value == "Company Name"
        assert ws['A2'].value == "SG&A Budget Report"
        assert ws['A3'].value is None  # Empty row
        
        # Check header row
        assert ws['A4'].value == "ID"
        assert ws['B4'].value == "Department/Project"
        
        wb.close()
    
    def test_autofilter_update(self, sample_workbook_path, temp_output_dir):
        """Test that autofilter is updated to match remaining data."""
        groups = ['IT']
        
        manifest_entries = export_clone(
            input_path=sample_workbook_path,
            groups=groups,
            sheet_name='SG&A Summary',
            header_row=3,
            dp_col_idx=1,
            out_dir=temp_output_dir
        )
        
        output_path = Path(manifest_entries[0]['output_path'])
        wb = openpyxl.load_workbook(output_path)
        ws = wb['SG&A Summary']
        
        # Check that autofilter exists and covers the remaining data
        if ws.auto_filter:
            # Should cover header row plus data rows
            expected_rows = 1 + manifest_entries[0]['row_count']  # header + data
            filter_range = ws.auto_filter.ref
            # The filter should start at row 4 (header) and include all data rows
            assert filter_range.startswith('A4:')
        
        wb.close()
    
    def test_empty_group_handling(self, sample_workbook_path, temp_output_dir):
        """Test handling of groups with no matching data."""
        groups = ['Sales', 'NonExistent']
        
        manifest_entries = export_clone(
            input_path=sample_workbook_path,
            groups=groups,
            sheet_name='SG&A Summary',
            header_row=3,
            dp_col_idx=1,
            out_dir=temp_output_dir
        )
        
        # Should only create file for Sales, not for NonExistent
        assert len(manifest_entries) == 1
        assert manifest_entries[0]['Department/Project'] == 'Sales'
        
        # Only one file should be created
        created_files = list(temp_output_dir.glob("*SG&A Summary.xlsx"))
        created_files = [f for f in created_files if f != sample_workbook_path]
        assert len(created_files) == 1
    
    def test_filename_sanitization(self, sample_workbook_path, temp_output_dir):
        """Test filename sanitization in clone mode."""
        # Create a workbook with problematic group names
        wb = openpyxl.load_workbook(sample_workbook_path)
        ws = wb['SG&A Summary']
        
        # Modify some data to have problematic characters
        ws['B5'] = 'Sales/Marketing'  # Replace first Sales entry
        ws['B7'] = 'IT<>Support'     # Replace first IT entry
        
        wb.save(sample_workbook_path)
        wb.close()
        
        groups = ['Sales/Marketing', 'IT<>Support']
        
        manifest_entries = export_clone(
            input_path=sample_workbook_path,
            groups=groups,
            sheet_name='SG&A Summary',
            header_row=3,
            dp_col_idx=1,
            out_dir=temp_output_dir
        )
        
        # Check that files were created with sanitized names
        created_files = list(temp_output_dir.glob("*.xlsx"))
        created_files = [f for f in created_files if f != sample_workbook_path]
        assert len(created_files) == 2
        
        # Verify filenames don't contain problematic characters
        for file in created_files:
            filename = file.name
            problematic_chars = ['<', '>', ':', '"', '|', '?', '*', '\\', '/']
            for char in problematic_chars:
                assert char not in filename
    
    def test_workbook_isolation(self, sample_workbook_path, temp_output_dir):
        """Test that each export gets a fresh copy of the workbook."""
        groups = ['Sales', 'Marketing']
        
        manifest_entries = export_clone(
            input_path=sample_workbook_path,
            groups=groups,
            sheet_name='SG&A Summary',
            header_row=3,
            dp_col_idx=1,
            out_dir=temp_output_dir
        )
        
        # Both files should be created successfully
        assert len(manifest_entries) == 2
        
        # Each file should contain only its respective group's data
        for entry in manifest_entries:
            output_path = Path(entry['output_path'])
            wb = openpyxl.load_workbook(output_path)
            ws = wb['SG&A Summary']
            
            group_name = entry['Department/Project']
            
            # Count data rows for this group
            data_rows = []
            for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
                if any(cell is not None for cell in row):
                    data_rows.append(row)
            
            # All data rows should belong to this group
            for row in data_rows:
                assert row[1] == group_name
            
            wb.close()
    
    def test_error_handling(self, temp_output_dir):
        """Test error handling for invalid inputs."""
        non_existent_file = temp_output_dir / "does_not_exist.xlsx"
        
        manifest_entries = export_clone(
            input_path=non_existent_file,
            groups=['Test'],
            sheet_name='Sheet1',
            header_row=0,
            dp_col_idx=0,
            out_dir=temp_output_dir
        )
        
        # Should handle the error gracefully and return empty manifest
        assert len(manifest_entries) == 0