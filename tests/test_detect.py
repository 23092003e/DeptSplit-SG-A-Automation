"""
Tests for sheet and header detection functionality.
"""

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from sga_splitter.detect import (
    find_target_sheet_name,
    detect_header_and_column,
    candidate_name_matches,
    _find_best_fuzzy_sheet
)


class TestFindTargetSheetName:
    """Test suite for find_target_sheet_name function."""
    
    def test_exact_match_found(self):
        """Test exact sheet name match."""
        wb = Workbook()
        wb.create_sheet("SG&A Summary Sheet")
        wb.create_sheet("Other Sheet")
        
        result = find_target_sheet_name(wb, "SG&A Summary Sheet", fuzzy=False)
        assert result == "SG&A Summary Sheet"
    
    def test_exact_match_not_found_fuzzy_disabled(self):
        """Test exact match not found with fuzzy disabled."""
        wb = Workbook()
        wb.create_sheet("Some Sheet")
        
        with pytest.raises(ValueError, match="Sheet 'Missing' not found"):
            find_target_sheet_name(wb, "Missing", fuzzy=False)
    
    def test_exact_match_not_found_fuzzy_enabled(self):
        """Test exact match not found with fuzzy enabled."""
        wb = Workbook()
        wb.create_sheet("SGA Summary Report")
        wb.create_sheet("Other Sheet")
        
        result = find_target_sheet_name(wb, "SG&A Summary", fuzzy=True)
        assert result == "SGA Summary Report"
    
    def test_no_request_fuzzy_enabled(self):
        """Test auto-detection with fuzzy matching."""
        wb = Workbook()
        wb.create_sheet("Data Sheet")
        wb.create_sheet("SG&A Summary Analysis")
        wb.create_sheet("Totals")
        
        result = find_target_sheet_name(wb, None, fuzzy=True)
        assert result == "SG&A Summary Analysis"
    
    def test_no_request_fuzzy_disabled(self):
        """Test fallback to first sheet when no request and fuzzy disabled."""
        wb = Workbook()
        first_sheet = wb.active
        first_sheet.title = "First Sheet"
        wb.create_sheet("Second Sheet")
        
        result = find_target_sheet_name(wb, None, fuzzy=False)
        assert result == "First Sheet"
    
    def test_empty_workbook(self):
        """Test error handling for empty workbook."""
        wb = Workbook()
        # Remove the default sheet
        wb.remove(wb.active)
        
        with pytest.raises(ValueError, match="Workbook contains no sheets"):
            find_target_sheet_name(wb, None, fuzzy=False)


class TestFindBestFuzzySheet:
    """Test suite for _find_best_fuzzy_sheet function."""
    
    def test_sga_summary_keywords(self):
        """Test detection based on SG&A and summary keywords."""
        sheet_names = ["Data", "SG&A Summary Report", "Totals"]
        result = _find_best_fuzzy_sheet(sheet_names)
        assert result == "SG&A Summary Report"
    
    def test_partial_keywords(self):
        """Test detection with partial keyword matches."""
        sheet_names = ["Revenue", "SGA Analysis", "Expenses"]
        result = _find_best_fuzzy_sheet(sheet_names)
        assert result == "SGA Analysis"
    
    def test_no_keywords_fallback(self):
        """Test fallback to first sheet when no keywords match."""
        sheet_names = ["Revenue", "Expenses", "Cash Flow"]
        result = _find_best_fuzzy_sheet(sheet_names)
        assert result == "Revenue"
    
    def test_case_insensitive_matching(self):
        """Test case-insensitive keyword matching."""
        sheet_names = ["data", "sg&a SUMMARY sheet", "other"]
        result = _find_best_fuzzy_sheet(sheet_names)
        assert result == "sg&a SUMMARY sheet"


class TestDetectHeaderAndColumn:
    """Test suite for detect_header_and_column function."""
    
    def create_test_worksheet(self, headers: list, header_row: int = 0) -> Worksheet:
        """Helper to create test worksheet with specific headers."""
        wb = Workbook()
        ws = wb.active
        
        # Add some data before header if needed
        for i in range(header_row):
            ws.cell(row=i + 1, column=1, value=f"Pre-header row {i + 1}")
        
        # Add headers
        for col_idx, header in enumerate(headers):
            ws.cell(row=header_row + 1, column=col_idx + 1, value=header)
        
        return ws
    
    def test_department_project_detection(self):
        """Test detection of Department/Project column."""
        headers = ["ID", "Department/Project", "Amount", "Description"]
        ws = self.create_test_worksheet(headers)
        
        header_row, dp_col = detect_header_and_column(ws)
        assert header_row == 0
        assert dp_col == 1
    
    def test_department_dash_project_detection(self):
        """Test detection of Department - Project column."""
        headers = ["ID", "Department - Project", "Amount"]
        ws = self.create_test_worksheet(headers)
        
        header_row, dp_col = detect_header_and_column(ws)
        assert header_row == 0
        assert dp_col == 1
    
    def test_dept_project_detection(self):
        """Test detection of Dept/Project column."""
        headers = ["Code", "Dept/Project", "Budget"]
        ws = self.create_test_worksheet(headers)
        
        header_row, dp_col = detect_header_and_column(ws)
        assert header_row == 0
        assert dp_col == 1
    
    def test_project_department_detection(self):
        """Test detection of Project/Department column."""
        headers = ["ID", "Project/Department", "Cost"]
        ws = self.create_test_worksheet(headers)
        
        header_row, dp_col = detect_header_and_column(ws)
        assert header_row == 0
        assert dp_col == 1
    
    def test_department_only_detection(self):
        """Test detection of standalone Department column."""
        headers = ["ID", "Department", "Amount"]
        ws = self.create_test_worksheet(headers)
        
        header_row, dp_col = detect_header_and_column(ws)
        assert header_row == 0
        assert dp_col == 1
    
    def test_project_only_detection(self):
        """Test detection of standalone Project column."""
        headers = ["Code", "Project", "Budget"]
        ws = self.create_test_worksheet(headers)
        
        header_row, dp_col = detect_header_and_column(ws)
        assert header_row == 0
        assert dp_col == 1
    
    def test_header_in_later_row(self):
        """Test detection when header is not in first row."""
        headers = ["", "", "Department/Project", "Amount"]
        ws = self.create_test_worksheet(headers, header_row=3)
        
        header_row, dp_col = detect_header_and_column(ws)
        assert header_row == 3
        assert dp_col == 2
    
    def test_case_insensitive_detection(self):
        """Test case-insensitive header detection."""
        headers = ["id", "DEPARTMENT/PROJECT", "amount"]
        ws = self.create_test_worksheet(headers)
        
        header_row, dp_col = detect_header_and_column(ws)
        assert header_row == 0
        assert dp_col == 1
    
    def test_whitespace_normalization(self):
        """Test detection with extra whitespace in headers."""
        headers = ["ID", "  Department / Project  ", "Amount"]
        ws = self.create_test_worksheet(headers)
        
        header_row, dp_col = detect_header_and_column(ws)
        assert header_row == 0
        assert dp_col == 1
    
    def test_no_matching_column_error(self):
        """Test error when no matching column is found."""
        headers = ["ID", "Name", "Amount", "Description"]
        ws = self.create_test_worksheet(headers)
        
        with pytest.raises(ValueError, match="Could not find Department/Project column"):
            detect_header_and_column(ws)


class TestCandidateNameMatches:
    """Test suite for candidate_name_matches function."""
    
    def test_exact_matches(self):
        """Test exact pattern matches."""
        test_cases = [
            "Department/Project",
            "Department-Project", 
            "Department - Project",
            "Project/Department",
            "Dept/Project",
            "Department",
            "Project",
            "Dept"
        ]
        
        for case in test_cases:
            assert candidate_name_matches(case), f"Should match: {case}"
    
    def test_case_insensitive(self):
        """Test case-insensitive matching."""
        test_cases = [
            "DEPARTMENT/PROJECT",
            "department/project",
            "Department/Project",
            "dEpArTmEnT/pRoJeCt"
        ]
        
        for case in test_cases:
            assert candidate_name_matches(case), f"Should match: {case}"
    
    def test_whitespace_handling(self):
        """Test whitespace normalization."""
        test_cases = [
            "  Department/Project  ",
            "Department / Project",
            "Department  /  Project",
            "\tDepartment/Project\n"
        ]
        
        for case in test_cases:
            assert candidate_name_matches(case), f"Should match: {case}"
    
    def test_non_matches(self):
        """Test strings that should not match."""
        test_cases = [
            "",
            "Name",
            "Amount", 
            "Department of Project",
            "Project Manager",
            "Department Name",
            "Project Code",
            "Budget Department",
            "Cost Center"
        ]
        
        for case in test_cases:
            assert not candidate_name_matches(case), f"Should not match: {case}"
    
    def test_empty_and_none(self):
        """Test empty string and None handling."""
        assert not candidate_name_matches("")
        assert not candidate_name_matches(None)