"""
Tests for multi-sheet processing functionality.
"""

import pytest
import pandas as pd
from pathlib import Path
import tempfile
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill

from sga_splitter.core import split_workbook_multi_sheet, _remove_unwanted_columns
from sga_splitter.exporters import export_clone_multi_sheet, _identify_columns_to_remove


class TestMultiSheetProcessing:
    """Test suite for multi-sheet processing functionality."""
    
    @pytest.fixture
    def temp_output_dir(self):
        """Create a temporary directory for test outputs."""
        temp_dir = Path(tempfile.mkdtemp())
        yield temp_dir
        # Cleanup after test
        shutil.rmtree(temp_dir)
    
    @pytest.fixture
    def multi_sheet_workbook_path(self, temp_output_dir):
        """Create a sample multi-sheet Excel workbook for testing."""
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Sheet 1 - Projects
        sheet1 = wb.create_sheet("Project Summary")
        sheet1['A1'] = "Company Budget Report"
        sheet1['A2'] = "Project Analysis"
        sheet1['A3'] = ""
        
        # Headers for sheet 1
        headers1 = ['ID', 'Project', 'Amount', 'Unnamed_1', 'Description']
        for col_idx, header in enumerate(headers1):
            cell = sheet1.cell(row=4, column=col_idx + 1, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Data for sheet 1
        data1 = [
            [1, 'Alpha', 1000, 'junk1', 'Project Alpha'],
            [2, 'Beta', 2000, 'junk2', 'Project Beta'],
            [3, 'Alpha', 1500, 'junk3', 'Project Alpha Phase 2'],
            [4, 'Gamma', 3000, 'junk4', 'Project Gamma']
        ]
        
        for row_idx, row_data in enumerate(data1):
            for col_idx, value in enumerate(row_data):
                sheet1.cell(row=5 + row_idx, column=col_idx + 1, value=value)
        
        # Create Sheet 2 - Departments
        sheet2 = wb.create_sheet("Department Budget")
        sheet2['A1'] = "Department Analysis"
        sheet2['A2'] = ""
        
        # Headers for sheet 2
        headers2 = ['Code', 'Department', 'Budget', 'Project/Department', 'Notes']
        for col_idx, header in enumerate(headers2):
            cell = sheet2.cell(row=3, column=col_idx + 1, value=header)
            cell.font = Font(bold=True)
        
        # Data for sheet 2
        data2 = [
            ['HR01', 'Human Resources', 5000, 'HR-Finance', 'HR expenses'],
            ['IT01', 'Information Technology', 8000, 'IT-Ops', 'IT infrastructure'],
            ['HR02', 'Human Resources', 3000, 'HR-Finance', 'HR training'],
            ['MK01', 'Marketing', 6000, 'MK-Sales', 'Marketing campaigns']
        ]
        
        for row_idx, row_data in enumerate(data2):
            for col_idx, value in enumerate(row_data):
                sheet2.cell(row=4 + row_idx, column=col_idx + 1, value=value)
        
        # Create Sheet 3 - More Departments
        sheet3 = wb.create_sheet("Dept Expenses")
        sheet3['A1'] = "Expense Report"
        
        # Headers for sheet 3
        headers3 = ['Item', 'Dept', 'Cost', 'Unnamed_Col']
        for col_idx, header in enumerate(headers3):
            cell = sheet3.cell(row=2, column=col_idx + 1, value=header)
            cell.font = Font(bold=True)
        
        # Data for sheet 3
        data3 = [
            ['Office Supplies', 'Operations', 500, 'temp'],
            ['Software Licenses', 'Information Technology', 1200, 'temp'],
            ['Training', 'Operations', 800, 'temp'],
            ['Equipment', 'Information Technology', 2000, 'temp']
        ]
        
        for row_idx, row_data in enumerate(data3):
            for col_idx, value in enumerate(row_data):
                sheet3.cell(row=3 + row_idx, column=col_idx + 1, value=value)
        
        # Save the workbook
        workbook_path = temp_output_dir / "test_multi_sheet.xlsx"
        wb.save(workbook_path)
        wb.close()
        
        return workbook_path
    
    def test_multi_sheet_basic_processing(self, multi_sheet_workbook_path, temp_output_dir):
        """Test basic multi-sheet processing functionality."""
        result = split_workbook_multi_sheet(
            input_path=multi_sheet_workbook_path,
            out_dir=temp_output_dir
        )
        
        # Check basic results
        assert result['total_files_created'] > 0
        assert len(result['sheets_processed']) == 3
        assert result['input_file'] == str(multi_sheet_workbook_path)
        
        # Check that sheet-specific directories were created
        sheet_dirs = list(temp_output_dir.glob("Sheet_*"))
        assert len(sheet_dirs) == 3
        
        # Verify files were created
        all_files = list(temp_output_dir.glob("**/*.xlsx"))
        input_file = multi_sheet_workbook_path
        created_files = [f for f in all_files if f != input_file]
        assert len(created_files) > 0
    
    def test_column_removal(self, multi_sheet_workbook_path, temp_output_dir):
        """Test that unwanted columns are removed from output."""
        result = split_workbook_multi_sheet(
            input_path=multi_sheet_workbook_path,
            out_dir=temp_output_dir,
            remove_columns=["unnamed", "project/department"]
        )
        
        # Check that files were created
        assert result['total_files_created'] > 0
        
        # Check one of the created files to verify column removal
        created_files = list(temp_output_dir.glob("**/*.xlsx"))
        input_file = multi_sheet_workbook_path
        output_files = [f for f in created_files if f != input_file]
        
        if output_files:
            # Open first output file and check columns
            test_file = output_files[0]
            wb = openpyxl.load_workbook(test_file)
            ws = wb.active
            
            # Get header row values
            header_values = []
            for cell in ws[1]:  # Assuming header is in first row of output
                if cell.value:
                    header_values.append(str(cell.value).lower())
            
            # Verify unwanted columns are not present
            unwanted_patterns = ['unnamed', 'project/department', 'department/project']
            for pattern in unwanted_patterns:
                assert not any(pattern in header for header in header_values), \
                    f"Found unwanted column pattern '{pattern}' in headers: {header_values}"
            
            wb.close()
    
    def test_sheet_specific_splitting(self, multi_sheet_workbook_path, temp_output_dir):
        """Test that sheets are split by correct criteria."""
        result = split_workbook_multi_sheet(
            input_path=multi_sheet_workbook_path,
            out_dir=temp_output_dir
        )
        
        # Verify sheet processing details
        sheets_processed = result['sheets_processed']
        assert len(sheets_processed) == 3
        
        # Check first sheet splits by project
        sheet1_info = sheets_processed[0]
        assert 'project' in sheet1_info['split_by'].lower()
        
        # Check second and third sheets split by department
        sheet2_info = sheets_processed[1]
        sheet3_info = sheets_processed[2]
        assert 'department' in sheet2_info['split_by'].lower()
        assert 'department' in sheet3_info['split_by'].lower()
    
    def test_formatting_preservation(self, multi_sheet_workbook_path, temp_output_dir):
        """Test that formatting is preserved in output files."""
        result = split_workbook_multi_sheet(
            input_path=multi_sheet_workbook_path,
            out_dir=temp_output_dir
        )
        
        # Get an output file to test
        created_files = list(temp_output_dir.glob("**/*.xlsx"))
        input_file = multi_sheet_workbook_path
        output_files = [f for f in created_files if f != input_file]
        
        if output_files:
            # Open first output file and check formatting
            test_file = output_files[0]
            wb = openpyxl.load_workbook(test_file)
            ws = wb.active
            
            # Find header row (look for bold formatting)
            header_row = None
            for row_idx in range(1, min(10, ws.max_row + 1)):
                for cell in ws[row_idx]:
                    if cell.font and cell.font.bold:
                        header_row = row_idx
                        break
                if header_row:
                    break
            
            assert header_row is not None, "Could not find formatted header row"
            
            # Verify header formatting is preserved
            header_cells = ws[header_row]
            bold_found = False
            for cell in header_cells:
                if cell.value and cell.font and cell.font.bold:
                    bold_found = True
                    break
            
            assert bold_found, "Header formatting not preserved"
            wb.close()


class TestColumnRemoval:
    """Test suite for column removal functionality."""
    
    def test_remove_unwanted_columns_dataframe(self):
        """Test removing unwanted columns from DataFrame."""
        df = pd.DataFrame({
            'ID': [1, 2, 3],
            'Project': ['A', 'B', 'C'],
            'Amount': [100, 200, 300],
            'Unnamed: 4': ['junk1', 'junk2', 'junk3'],
            'Project/Department': ['P1', 'P2', 'P3'],
            'Description': ['Desc1', 'Desc2', 'Desc3']
        })
        
        remove_patterns = ['unnamed', 'project/department']
        result_df = _remove_unwanted_columns(df, remove_patterns)
        
        # Check that unwanted columns are removed
        expected_columns = ['ID', 'Project', 'Amount', 'Description']
        assert list(result_df.columns) == expected_columns
        
        # Check that data is preserved
        assert len(result_df) == 3
        assert result_df['Project'].tolist() == ['A', 'B', 'C']
    
    def test_identify_columns_to_remove_worksheet(self):
        """Test identifying columns to remove from worksheet."""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Create header row
        headers = ['ID', 'Project', 'Unnamed: 3', 'Project/Department', 'Notes']
        for col_idx, header in enumerate(headers):
            ws.cell(row=1, column=col_idx + 1, value=header)
        
        remove_patterns = ['unnamed', 'project/department']
        columns_to_remove = _identify_columns_to_remove(ws, 0, remove_patterns)
        
        # Should identify columns 2 and 3 for removal (0-based indexing)
        expected_removals = [2, 3]  # 'Unnamed: 3' and 'Project/Department'
        assert columns_to_remove == expected_removals
        
        wb.close()
    
    def test_preserve_split_column(self):
        """Test that the split column is preserved even if it matches removal patterns."""
        df = pd.DataFrame({
            'ID': [1, 2, 3],
            'Project/Department': ['A', 'B', 'C'],  # Split column that matches removal pattern
            'Amount': [100, 200, 300],
            'Unnamed: 4': ['junk1', 'junk2', 'junk3']
        })
        
        remove_patterns = ['unnamed', 'project/department']
        # Preserve the split column even though it matches removal pattern
        result_df = _remove_unwanted_columns(df, remove_patterns, preserve_column='Project/Department')
        
        # Split column should be preserved, unwanted columns removed
        expected_columns = ['ID', 'Project/Department', 'Amount']
        assert list(result_df.columns) == expected_columns
        assert len(result_df) == 3
        assert result_df['Project/Department'].tolist() == ['A', 'B', 'C']
    
    def test_no_column_removal(self):
        """Test that no columns are removed when patterns don't match."""
        df = pd.DataFrame({
            'ID': [1, 2, 3],
            'Project': ['A', 'B', 'C'],
            'Amount': [100, 200, 300]
        })
        
        remove_patterns = ['nonexistent']
        result_df = _remove_unwanted_columns(df, remove_patterns)
        
        # All columns should be preserved
        assert list(result_df.columns) == list(df.columns)
        assert len(result_df) == len(df)


class TestMultiSheetExport:
    """Test suite for multi-sheet export functionality."""
    
    @pytest.fixture
    def temp_output_dir(self):
        """Create a temporary directory for test outputs."""
        temp_dir = Path(tempfile.mkdtemp())
        yield temp_dir
        shutil.rmtree(temp_dir)
    
    @pytest.fixture
    def simple_workbook_path(self, temp_output_dir):
        """Create a simple workbook for export testing."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Sheet"
        
        # Headers
        headers = ['ID', 'Department', 'Amount', 'Unnamed_Col']
        for col_idx, header in enumerate(headers):
            ws.cell(row=1, column=col_idx + 1, value=header)
        
        # Data
        data = [
            [1, 'IT', 1000, 'junk1'],
            [2, 'HR', 2000, 'junk2'],
            [3, 'IT', 1500, 'junk3']
        ]
        
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                ws.cell(row=2 + row_idx, column=col_idx + 1, value=value)
        
        workbook_path = temp_output_dir / "simple_test.xlsx"
        wb.save(workbook_path)
        wb.close()
        return workbook_path
    
    def test_export_clone_multi_sheet(self, simple_workbook_path, temp_output_dir):
        """Test the export_clone_multi_sheet function."""
        groups = ['IT', 'HR']
        sheet_name = 'Test Sheet'
        header_row = 0
        split_col_idx = 1  # Department column
        remove_columns = ['unnamed']
        original_split_col_idx = 1
        
        manifest_entries = export_clone_multi_sheet(
            input_path=simple_workbook_path,
            groups=groups,
            sheet_name=sheet_name,
            header_row=header_row,
            split_col_idx=split_col_idx,
            out_dir=temp_output_dir,
            remove_columns=remove_columns,
            original_split_col_idx=original_split_col_idx
        )
        
        # Check manifest entries
        assert len(manifest_entries) == 2
        
        for entry in manifest_entries:
            assert entry['mode'] == 'clone_multi'
            assert 'Group' in entry
            assert 'output_path' in entry
            assert entry['row_count'] > 0
        
        # Verify files were created
        created_files = list(temp_output_dir.glob("*.xlsx"))
        # Filter out input file
        output_files = [f for f in created_files if f != simple_workbook_path]
        assert len(output_files) == 2
        
        # Check that unwanted columns were removed
        for output_file in output_files:
            wb = openpyxl.load_workbook(output_file)
            ws = wb.active
            
            # Get header values
            header_values = [str(ws.cell(row=1, column=col).value or "") 
                           for col in range(1, ws.max_column + 1)]
            
            # Should not contain 'Unnamed_Col'
            assert 'Unnamed_Col' not in header_values
            
            wb.close()