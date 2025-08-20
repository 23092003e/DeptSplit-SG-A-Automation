"""
Export utilities for generating split Excel files in fast and clone modes.
"""

from pathlib import Path
from typing import List, Dict, Any
import logging

import pandas as pd
import xlsxwriter #type: ignore
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .io_utils import sanitize_filename, generate_unique_filename

logger = logging.getLogger(__name__)


def export_fast(
    df: pd.DataFrame,
    groups: List[str],
    dp_col: str,
    sheet_name: str,
    out_dir: Path
) -> List[Dict[str, Any]]:
    """
    Export groups using fast mode (pandas + xlsxwriter).
    
    Args:
        df: Source DataFrame
        groups: List of Department/Project values to export
        dp_col: Name of the Department/Project column
        sheet_name: Name for the output sheet
        out_dir: Output directory
        
    Returns:
        List of manifest entries for created files
    """
    manifest_entries = []
    
    for group in groups:
        # Filter data for this group
        group_df = df[df[dp_col] == group].copy()
        
        if group_df.empty:
            logger.warning(f"No data found for group: {group}")
            continue
        
        # Generate output filename
        sanitized_name = sanitize_filename(group)
        base_filename = f"{sanitized_name} - SG&A Summary"
        output_path = generate_unique_filename(out_dir / base_filename, ".xlsx")
        
        try:
            # Write using xlsxwriter for performance
            with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
                group_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get the workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Freeze the header row
                worksheet.freeze_panes(1, 0)
                
                # Add autofilter
                if not group_df.empty:
                    last_col = xlsxwriter.utility.xl_col_to_name(len(group_df.columns) - 1)
                    worksheet.autofilter(f'A1:{last_col}{len(group_df) + 1}')
                
                # Format header row
                header_format = workbook.add_format({
                    'bold': True,
                    'bg_color': '#D9E1F2',
                    'border': 1
                })
                
                for col_idx, col_name in enumerate(group_df.columns):
                    worksheet.write(0, col_idx, col_name, header_format)
                    # Auto-adjust column width
                    worksheet.set_column(col_idx, col_idx, len(str(col_name)) + 2)
            
            manifest_entries.append({
                'Department/Project': group,
                'output_path': str(output_path),
                'row_count': len(group_df),
                'mode': 'fast'
            })
            
            logger.info(f"Created {output_path} with {len(group_df)} rows")
            
        except Exception as e:
            logger.error(f"Failed to create file for group '{group}': {e}")
            continue
    
    return manifest_entries


def export_clone(
    input_path: Path,
    groups: List[str],
    sheet_name: str,
    header_row: int,
    dp_col_idx: int,
    out_dir: Path
) -> List[Dict[str, Any]]:
    """
    Export groups using clone mode (openpyxl with style preservation).
    
    Args:
        input_path: Path to source Excel file
        groups: List of Department/Project values to export
        sheet_name: Name of the sheet to clone
        header_row: 0-based index of header row
        dp_col_idx: 0-based index of Department/Project column
        out_dir: Output directory
        
    Returns:
        List of manifest entries for created files
    """
    manifest_entries = []
    
    for group in groups:
        try:
            # Load fresh copy of workbook for each group
            wb = openpyxl.load_workbook(input_path)
            
            if sheet_name not in wb.sheetnames:
                logger.error(f"Sheet '{sheet_name}' not found in workbook")
                continue
            
            ws = wb[sheet_name]
            
            # Find rows to keep (header + matching data rows)
            rows_to_keep = set()
            rows_to_keep.update(range(1, header_row + 2))  # Keep everything up to and including header
            
            row_count = 0
            for row_idx in range(header_row + 2, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=dp_col_idx + 1).value
                if cell_value is not None and str(cell_value).strip() == group:
                    rows_to_keep.add(row_idx)
                    row_count += 1
            
            if row_count == 0:
                logger.warning(f"No data rows found for group: {group}")
                wb.close()
                continue
            
            # Delete rows that don't match (from bottom to top to preserve indices)
            rows_to_delete = []
            for row_idx in range(header_row + 2, ws.max_row + 1):
                if row_idx not in rows_to_keep:
                    rows_to_delete.append(row_idx)
            
            # Delete from bottom to top
            for row_idx in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(row_idx)
            
            # Update autofilter if it exists
            if ws.auto_filter:
                ws.auto_filter.ref = None
            
            # Reapply autofilter to the remaining data
            if ws.max_row > header_row + 1:  # If we have data rows
                end_col = openpyxl.utils.get_column_letter(ws.max_column)
                ws.auto_filter.ref = f"A{header_row + 1}:{end_col}{ws.max_row}"
            
            # Generate output filename
            sanitized_name = sanitize_filename(group)
            base_filename = f"{sanitized_name} - SG&A Summary"
            output_path = generate_unique_filename(out_dir / base_filename, ".xlsx")
            
            # Save the modified workbook
            wb.save(output_path)
            wb.close()
            
            manifest_entries.append({
                'Department/Project': group,
                'output_path': str(output_path),
                'row_count': row_count,
                'mode': 'clone'
            })
            
            logger.info(f"Created {output_path} with {row_count} rows (clone mode)")
            
        except Exception as e:
            logger.error(f"Failed to create file for group '{group}' in clone mode: {e}")
            continue
    
    return manifest_entries


def _should_skip_row_for_group(ws: Worksheet, row_idx: int, dp_col_idx: int, target_group: str) -> bool:
    """
    Check if a row should be skipped (doesn't match the target group).
    
    Args:
        ws: Worksheet to check
        row_idx: 1-based row index
        dp_col_idx: 0-based column index for Department/Project
        target_group: Target group value to match
        
    Returns:
        True if row should be skipped
    """
    cell_value = ws.cell(row=row_idx, column=dp_col_idx + 1).value
    
    if cell_value is None:
        return True
    
    return str(cell_value).strip() != target_group


def export_clone_multi_sheet(
    input_path: Path,
    groups: List[str],
    sheet_name: str,
    header_row: int,
    split_col_idx: int,
    out_dir: Path,
    remove_columns: List[str],
    original_split_col_idx: int
) -> List[Dict[str, Any]]:
    """
    Export groups using clone mode with column removal and enhanced formatting preservation.
    
    Args:
        input_path: Path to source Excel file
        groups: List of values to export
        sheet_name: Name of the sheet to clone
        header_row: 0-based index of header row
        split_col_idx: 0-based index of split column (after column removal)
        out_dir: Output directory
        remove_columns: List of column patterns to remove
        original_split_col_idx: Original split column index before column removal
        
    Returns:
        List of manifest entries for created files
    """
    manifest_entries = []
    
    for group in groups:
        try:
            # Load fresh copy of workbook for each group
            wb = openpyxl.load_workbook(input_path)
            
            if sheet_name not in wb.sheetnames:
                logger.error(f"Sheet '{sheet_name}' not found in workbook")
                continue
            
            ws = wb[sheet_name]
            
            # Remove unwanted columns first (but preserve the split column)
            columns_to_remove = _identify_columns_to_remove(ws, header_row, remove_columns, preserve_col_idx=original_split_col_idx)
            
            # Adjust split column index if columns before it are removed
            adjusted_split_col_idx = original_split_col_idx
            for col_idx in sorted(columns_to_remove):
                if col_idx < original_split_col_idx:
                    adjusted_split_col_idx -= 1
            
            # Remove columns from right to left to preserve indices
            for col_idx in sorted(columns_to_remove, reverse=True):
                ws.delete_cols(col_idx + 1)  # openpyxl uses 1-based indexing
            
            # Find rows to keep (header + matching data rows)
            rows_to_keep = set()
            rows_to_keep.update(range(1, header_row + 2))  # Keep everything up to and including header
            
            row_count = 0
            for row_idx in range(header_row + 2, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=adjusted_split_col_idx + 1).value
                if cell_value is not None and str(cell_value).strip() == group:
                    rows_to_keep.add(row_idx)
                    row_count += 1
            
            if row_count == 0:
                logger.warning(f"No data rows found for group: {group}")
                wb.close()
                continue
            
            # Delete rows that don't match (from bottom to top to preserve indices)
            rows_to_delete = []
            for row_idx in range(header_row + 2, ws.max_row + 1):
                if row_idx not in rows_to_keep:
                    rows_to_delete.append(row_idx)
            
            # Delete from bottom to top
            for row_idx in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(row_idx)
            
            # Preserve and enhance formatting
            _preserve_formatting(ws, header_row)
            
            # Update autofilter if it exists
            if ws.auto_filter:
                ws.auto_filter.ref = None
            
            # Reapply autofilter to the remaining data
            if ws.max_row > header_row + 1:  # If we have data rows
                end_col = openpyxl.utils.get_column_letter(ws.max_column)
                ws.auto_filter.ref = f"A{header_row + 1}:{end_col}{ws.max_row}"
            
            # Generate output filename
            sanitized_name = sanitize_filename(group)
            base_filename = f"{sanitized_name} - {sheet_name}"
            output_path = generate_unique_filename(out_dir / base_filename, ".xlsx")
            
            # Save the modified workbook
            wb.save(output_path)
            wb.close()
            
            manifest_entries.append({
                'Sheet': sheet_name,
                'Group': group,
                'output_path': str(output_path),
                'row_count': row_count,
                'mode': 'clone_multi'
            })
            
            logger.info(f"Created {output_path} with {row_count} rows (multi-sheet clone mode)")
            
        except Exception as e:
            logger.error(f"Failed to create file for group '{group}' in multi-sheet clone mode: {e}")
            continue
    
    return manifest_entries


def _identify_columns_to_remove(ws: Worksheet, header_row: int, remove_patterns: List[str], preserve_col_idx: int = None) -> List[int]:
    """
    Identify column indices to remove based on patterns, but preserve specified column.
    
    Args:
        ws: Worksheet to analyze
        header_row: 0-based header row index
        remove_patterns: List of patterns to match for removal
        preserve_col_idx: 0-based column index to preserve even if it matches patterns
        
    Returns:
        List of 0-based column indices to remove
    """
    columns_to_remove = []
    
    if not remove_patterns:
        return columns_to_remove
    
    # Get header row values
    max_col = ws.max_column
    for col_idx in range(max_col):
        # Always preserve the split column
        if preserve_col_idx is not None and col_idx == preserve_col_idx:
            logger.info(f"Preserving split column at index {col_idx + 1}")
            continue
            
        header_cell = ws.cell(row=header_row + 1, column=col_idx + 1)
        header_value = str(header_cell.value or "").strip().lower()
        
        should_remove = False
        for pattern in remove_patterns:
            pattern_lower = pattern.lower()
            if (pattern_lower in header_value or 
                header_value.startswith('unnamed') or
                'project/department' in header_value or
                'department/project' in header_value):
                should_remove = True
                break
        
        if should_remove:
            columns_to_remove.append(col_idx)
            logger.info(f"Marking column {col_idx + 1} ('{header_value}') for removal")
    
    return columns_to_remove


def _preserve_formatting(ws: Worksheet, header_row: int):
    """
    Preserve and enhance worksheet formatting.
    
    Args:
        ws: Worksheet to format
        header_row: 0-based header row index
    """
    try:
        # Preserve column widths by copying from original
        # (This is already preserved by openpyxl when not explicitly changed)
        
        # Ensure header row has proper formatting
        for cell in ws[header_row + 1]:  # openpyxl uses 1-based row indexing
            if cell.value:
                # Keep existing formatting but ensure visibility
                if not cell.font.bold:
                    from openpyxl.styles import Font
                    cell.font = Font(bold=True)
        
        # Auto-adjust column widths if they seem too narrow
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            # Set a reasonable width (minimum 10, maximum 50)
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
            
    except Exception as e:
        logger.warning(f"Could not fully preserve formatting: {e}")
        # Continue processing even if formatting fails



def _write_dataframe_to_worksheet(df, ws):
    """
    Write DataFrame data to worksheet with proper formatting.
    
    Args:
        df: DataFrame to write
        ws: Worksheet to write to
    """
    # Write headers
    for col_idx, column_name in enumerate(df.columns):
        ws.cell(row=1, column=col_idx + 1, value=column_name)
    
    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row):
            ws.cell(row=row_idx, column=col_idx + 1, value=value)




def _apply_basic_formatting(ws):
    """
    Apply basic formatting to the worksheet.
    
    Args:
        ws: Worksheet to format
    """
    try:
        from openpyxl.styles import Font, PatternFill
        
        # Format header row
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        for cell in ws[1]:  # First row
            if cell.value:
                cell.font = header_font
                cell.fill = header_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
            
        # Add autofilter
        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}"
            
    except Exception as e:
        logger.warning(f"Could not apply basic formatting: {e}")


def create_multi_sheet_files_per_group(
    input_path: Path,
    groups: List[str],
    processed_sheets: List[Dict[str, Any]],
    out_dir: Path,
    remove_columns: List[str]
) -> List[Dict[str, Any]]:
    """
    Create one Excel file per group, each containing all three sheets filtered for that group.
    
    Args:
        input_path: Path to original input file
        groups: List of unique groups across all sheets
        processed_sheets: List of processed sheet data
        out_dir: Output directory
        remove_columns: Columns to remove
        
    Returns:
        List of manifest entries for created files
    """
    logger.info(f"Creating {len(groups)} files - one per group with all 3 sheets")
    
    manifest_entries = []
    
    # Load original workbook for copying formatting
    original_wb = openpyxl.load_workbook(input_path)
    
    try:
        for group in groups:
            logger.info(f"Creating file for group: {group}")
            
            # Create new workbook for this group
            group_wb = openpyxl.Workbook()
            group_wb.remove(group_wb.active)  # Remove default sheet
            
            sheets_created = []
            total_rows = 0
            
            try:
                for sheet_data in processed_sheets:
                    sheet_name = sheet_data['config']['sheet_name']
                    
                    # Check if this group exists in this sheet
                    if group not in sheet_data['groups']:
                        logger.info(f"Group '{group}' not found in sheet '{sheet_name}' - creating empty filtered sheet")
                    
                    # Get original sheet for copying formatting
                    original_ws = original_wb[sheet_name]
                    
                    # Create new sheet in group workbook
                    new_ws = group_wb.create_sheet(sheet_name)
                    
                    # Copy sheet data filtered for this group
                    rows_added = _copy_sheet_filtered_for_group(
                        original_ws=original_ws,
                        new_ws=new_ws,
                        group=group,
                        sheet_data=sheet_data,
                        remove_columns=remove_columns
                    )
                    
                    sheets_created.append(sheet_name)
                    total_rows += rows_added
                
                # Generate filename and save
                safe_group_name = sanitize_filename(group)
                output_filename = f"{safe_group_name}_SG&A YTD Budget report.xlsx"
                output_path = out_dir / output_filename
                output_path = generate_unique_filename(output_path)
                
                group_wb.save(output_path)
                logger.info(f"Created file for group '{group}': {output_path}")
                
                # Create manifest entry
                manifest_entry = {
                    'mode': 'multi_sheet_per_group',
                    'Group': group,
                    'output_path': str(output_path),
                    'sheets_included': ', '.join(sheets_created),
                    'total_rows': total_rows
                }
                manifest_entries.append(manifest_entry)
                
            finally:
                group_wb.close()
    
    finally:
        original_wb.close()
    
    return manifest_entries


def _copy_sheet_filtered_for_group(
    original_ws: Worksheet,
    new_ws: Worksheet,
    group: str,
    sheet_data: Dict[str, Any],
    remove_columns: List[str]
) -> int:
    # Get the columns to remove for this specific sheet
    columns_to_remove = _identify_columns_to_remove(
        original_ws, 
        sheet_data['header_row'], 
        remove_columns,
        preserve_col_idx=sheet_data['split_col_idx']
    )
    
    split_col_name = sheet_data['split_col_name']
    df = sheet_data['df']
    header_row_idx = sheet_data['header_row']
    
    # Filter data for this group
    group_df = df[df[split_col_name] == group].copy() if group in df[split_col_name].values else pd.DataFrame()
    
    logger.info(f"Group '{group}' has {len(group_df)} rows in DataFrame")
    
    rows_added = 0
    
    # FIXED: Create proper mapping between original columns, cleaned DataFrame, and output
    df_to_output_mapping = []
    output_col_idx = 1
    df_col_idx = 0  # Track position in cleaned DataFrame
    
    # Build header mapping to match DataFrame columns with original positions
    header_row = header_row_idx + 1  # Convert to 1-based
    original_headers = []
    for col_idx in range(1, original_ws.max_column + 1):
        header_cell = original_ws.cell(row=header_row, column=col_idx)
        original_headers.append(str(header_cell.value or "").strip())
    
    # Create mapping for non-removed columns
    for original_col_idx in range(1, original_ws.max_column + 1):
        # Skip columns that should be removed
        if (original_col_idx - 1) in columns_to_remove:
            continue
        
        # Find corresponding column in cleaned DataFrame
        original_header = original_headers[original_col_idx - 1]
        
        # Find matching column in DataFrame by header name
        matching_df_col_idx = None
        for i, df_col_name in enumerate(df.columns):
            if str(df_col_name).strip() == original_header:
                matching_df_col_idx = i
                break
        
        # If exact match not found, use positional mapping
        if matching_df_col_idx is None and df_col_idx < len(df.columns):
            matching_df_col_idx = df_col_idx
        
        df_to_output_mapping.append({
            'original_col': original_col_idx,
            'output_col': output_col_idx,
            'df_col_idx': matching_df_col_idx,
            'original_header': original_header
        })
        
        output_col_idx += 1
        df_col_idx += 1
    
    # Rest of the function remains the same...
    # Copy structure rows (everything up to and including header)
    for row_idx in range(1, header_row_idx + 2):
        for mapping in df_to_output_mapping:
            original_col = mapping['original_col']
            output_col = mapping['output_col']
            
            original_cell = original_ws.cell(row=row_idx, column=original_col)
            new_cell = new_ws.cell(row=row_idx, column=output_col)
            
            new_cell.value = original_cell.value
            _copy_cell_formatting(original_cell, new_cell)
    
    # Add filtered data rows with FIXED column mapping
    if not group_df.empty:
        data_start_row = header_row_idx + 2
        
        for df_row_idx, (_, row) in enumerate(group_df.iterrows()):
            excel_row_idx = data_start_row + df_row_idx
            
            for mapping in df_to_output_mapping:
                output_col = mapping['output_col']
                df_col_idx = mapping['df_col_idx']
                
                # Get value from DataFrame using correct column index
                if df_col_idx is not None and df_col_idx < len(df.columns):
                    col_name = df.columns[df_col_idx]
                    value = row.get(col_name, '')
                else:
                    value = ''
                
                new_cell = new_ws.cell(row=excel_row_idx, column=output_col)
                new_cell.value = value
                
                # Copy formatting from template
                original_data_row = header_row_idx + 2
                if original_data_row <= original_ws.max_row:
                    template_cell = original_ws.cell(row=original_data_row, column=mapping['original_col'])
                    _copy_cell_formatting(template_cell, new_cell)
            
            rows_added += 1
    
    # Copy column widths
    for mapping in df_to_output_mapping:
        original_col = mapping['original_col']
        output_col = mapping['output_col']
        
        original_letter = openpyxl.utils.get_column_letter(original_col)
        output_letter = openpyxl.utils.get_column_letter(output_col)
        
        if original_letter in original_ws.column_dimensions:
            new_ws.column_dimensions[output_letter].width = original_ws.column_dimensions[original_letter].width
    
    # Copy row heights for structure rows
    for row_idx in range(1, header_row_idx + 2):
        if row_idx in original_ws.row_dimensions:
            new_ws.row_dimensions[row_idx].height = original_ws.row_dimensions[row_idx].height
    
    logger.info(f"Copied sheet '{new_ws.title}' for group '{group}' with {rows_added} data rows")
    return rows_added


def _copy_cell_formatting(original_cell, new_cell):
    """Helper function to copy cell formatting safely."""
    try:
        if original_cell.font:
            from openpyxl.styles import Font
            new_cell.font = Font(
                name=original_cell.font.name,
                size=original_cell.font.size,
                bold=original_cell.font.bold,
                italic=original_cell.font.italic,
                color=original_cell.font.color
            )
        if original_cell.fill:
            from openpyxl.styles import PatternFill
            if hasattr(original_cell.fill, 'patternType') and original_cell.fill.patternType:
                new_cell.fill = PatternFill(
                    patternType=original_cell.fill.patternType,
                    start_color=original_cell.fill.start_color,
                    end_color=original_cell.fill.end_color
                )
        if original_cell.border:
            from openpyxl.styles import Border
            new_cell.border = Border(
                left=original_cell.border.left,
                right=original_cell.border.right,
                top=original_cell.border.top,
                bottom=original_cell.border.bottom
            )
        if original_cell.alignment:
            from openpyxl.styles import Alignment
            new_cell.alignment = Alignment(
                horizontal=original_cell.alignment.horizontal,
                vertical=original_cell.alignment.vertical,
                wrap_text=original_cell.alignment.wrap_text
            )
        if original_cell.number_format:
            new_cell.number_format = original_cell.number_format
    except Exception as e:
        logger.debug(f"Could not copy formatting: {e}")


def _copy_sheet_with_processed_data(
    original_ws: Worksheet,
    new_ws: Worksheet,
    sheet_data: Dict[str, Any],
    remove_columns: List[str]
) -> None:
    """
    Copy original sheet structure with processed data (columns removed).
    
    Args:
        original_ws: Original worksheet
        new_ws: New worksheet to populate
        sheet_data: Processed sheet data
        remove_columns: Columns to remove
    """
    # Get the columns to remove for this specific sheet
    columns_to_remove = _identify_columns_to_remove(
        original_ws, 
        sheet_data['header_row'], 
        remove_columns,
        preserve_col_idx=sheet_data['split_col_idx']
    )
    
    # Copy rows from original, removing unwanted columns
    for row_idx in range(1, original_ws.max_row + 1):
        new_col_idx = 1
        
        for col_idx in range(1, original_ws.max_column + 1):
            # Skip columns that should be removed
            if (col_idx - 1) in columns_to_remove:
                continue
                
            # Get original cell
            original_cell = original_ws.cell(row=row_idx, column=col_idx)
            
            # Create new cell
            new_cell = new_ws.cell(row=row_idx, column=new_col_idx)
            
            # Copy value and formatting
            new_cell.value = original_cell.value
            try:
                if original_cell.font:
                    from openpyxl.styles import Font
                    new_cell.font = Font(
                        name=original_cell.font.name,
                        size=original_cell.font.size,
                        bold=original_cell.font.bold,
                        italic=original_cell.font.italic,
                        color=original_cell.font.color
                    )
                if original_cell.fill:
                    from openpyxl.styles import PatternFill
                    if hasattr(original_cell.fill, 'patternType') and original_cell.fill.patternType:
                        new_cell.fill = PatternFill(
                            patternType=original_cell.fill.patternType,
                            start_color=original_cell.fill.start_color,
                            end_color=original_cell.fill.end_color
                        )
                if original_cell.border:
                    from openpyxl.styles import Border
                    new_cell.border = Border(
                        left=original_cell.border.left,
                        right=original_cell.border.right,
                        top=original_cell.border.top,
                        bottom=original_cell.border.bottom
                    )
                if original_cell.alignment:
                    from openpyxl.styles import Alignment
                    new_cell.alignment = Alignment(
                        horizontal=original_cell.alignment.horizontal,
                        vertical=original_cell.alignment.vertical,
                        wrap_text=original_cell.alignment.wrap_text
                    )
                if original_cell.number_format:
                    new_cell.number_format = original_cell.number_format
            except Exception as e:
                logger.debug(f"Could not copy formatting for cell {row_idx},{col_idx}: {e}")
                
            new_col_idx += 1
    
    # Copy column widths (adjust for removed columns)
    new_col_idx = 1
    for col_idx in range(1, original_ws.max_column + 1):
        if (col_idx - 1) not in columns_to_remove:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            new_col_letter = openpyxl.utils.get_column_letter(new_col_idx)
            
            if col_letter in original_ws.column_dimensions:
                new_ws.column_dimensions[new_col_letter].width = original_ws.column_dimensions[col_letter].width
            
            new_col_idx += 1
    
    # Copy row heights
    for row_idx in range(1, original_ws.max_row + 1):
        if row_idx in original_ws.row_dimensions:
            new_ws.row_dimensions[row_idx].height = original_ws.row_dimensions[row_idx].height
    
    logger.info(f"Copied sheet data with {len(columns_to_remove)} columns removed")