"""
Sheet and header detection utilities for SG&A splitter.
"""

import re
from typing import Optional, Tuple
from difflib import SequenceMatcher

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def find_target_sheet_name(wb: Workbook, requested: Optional[str], fuzzy: bool) -> str:
    """
    Find the target sheet name, with optional fuzzy matching.
    
    Args:
        wb: The openpyxl workbook
        requested: The explicitly requested sheet name, or None
        fuzzy: Whether to enable fuzzy matching
        
    Returns:
        The sheet name to use
        
    Raises:
        ValueError: If no suitable sheet is found
    """
    sheet_names = wb.sheetnames
    
    if not sheet_names:
        raise ValueError("Workbook contains no sheets")
    
    # If no specific sheet requested, try fuzzy matching if enabled
    if requested is None:
        if fuzzy:
            return _find_best_fuzzy_sheet(sheet_names)
        else:
            # Default to first sheet with warning
            return sheet_names[0]
    
    # Try exact match first
    if requested in sheet_names:
        return requested
    
    # If fuzzy enabled and exact match failed, try fuzzy
    if fuzzy:
        return _find_best_fuzzy_sheet(sheet_names, requested)
    
    raise ValueError(f"Sheet '{requested}' not found. Available sheets: {sheet_names}")


def _find_best_fuzzy_sheet(sheet_names: list[str], requested: Optional[str] = None) -> str:
    """
    Find the best fuzzy match for SG&A summary sheet.
    
    Args:
        sheet_names: List of available sheet names
        requested: Optional specific sheet name to match against
        
    Returns:
        Best matching sheet name
        
    Raises:
        ValueError: If no suitable fuzzy match found
    """
    if requested:
        # Find best match for the requested name
        matches = [(name, SequenceMatcher(None, requested.lower(), name.lower()).ratio()) 
                   for name in sheet_names]
        matches.sort(key=lambda x: x[1], reverse=True)
        if matches[0][1] > 0.6:  # Minimum similarity threshold
            return matches[0][0]
    
    # Look for sheets containing SG&A and summary keywords
    sga_keywords = ["sg&a", "sga", "selling", "general", "administrative"]
    summary_keywords = ["summary", "sum", "total", "consolidated"]
    
    scored_sheets = []
    for name in sheet_names:
        name_lower = name.lower()
        sga_score = sum(1 for kw in sga_keywords if kw in name_lower)
        summary_score = sum(1 for kw in summary_keywords if kw in name_lower)
        total_score = sga_score * 2 + summary_score  # Weight SG&A keywords higher
        
        if total_score > 0:
            scored_sheets.append((name, total_score))
    
    if scored_sheets:
        scored_sheets.sort(key=lambda x: x[1], reverse=True)
        return scored_sheets[0][0]
    
    # Fallback to first sheet
    return sheet_names[0]


def detect_header_and_column(ws: Worksheet) -> Tuple[int, int]:
    """
    Detect the header row and Department/Project column.
    
    Args:
        ws: The worksheet to analyze
        
    Returns:
        Tuple of (header_row_index, dp_column_index) (0-based)
        
    Raises:
        ValueError: If no suitable header/column found
    """
    max_rows_to_scan = min(50, ws.max_row)
    
    for row_idx in range(max_rows_to_scan):
        row_values = []
        for col_idx in range(min(20, ws.max_column)):  # Scan first 20 columns
            cell = ws.cell(row=row_idx + 1, column=col_idx + 1)
            value = str(cell.value or "").strip()
            row_values.append(value)
        
        # Check if any column header matches Department/Project pattern
        for col_idx, header in enumerate(row_values):
            if candidate_name_matches(header):
                return row_idx, col_idx
    
    # If we get here, no matching header was found
    sample_headers = []
    if max_rows_to_scan > 0:
        first_row_values = []
        for col_idx in range(min(5, ws.max_column)):
            cell = ws.cell(row=1, column=col_idx + 1)
            value = str(cell.value or "").strip()
            if value:
                first_row_values.append(value)
        sample_headers = first_row_values[:5]
    
    raise ValueError(
        f"Could not find Department/Project column. "
        f"Sample headers from first row: {sample_headers}"
    )


def candidate_name_matches(header: str) -> bool:
    """
    Check if a header name matches Department/Project column patterns.
    
    Args:
        header: The header text to check
        
    Returns:
        True if it matches expected patterns
    """
    if not header:
        return False
    
    # Normalize the header: remove extra whitespace, lowercase
    normalized = re.sub(r'\s+', ' ', header.strip().lower())
    
    # Define patterns to match
    patterns = [
        r'^department\s*[/-]?\s*project$',
        r'^project\s*[/-]?\s*department$',
        r'^dept\s*[/-]?\s*project$',
        r'^project\s*[/-]?\s*dept$',
        r'^department$',
        r'^project$',
        r'^dept$',
    ]
    
    for pattern in patterns:
        if re.match(pattern, normalized):
            return True
    
    return False